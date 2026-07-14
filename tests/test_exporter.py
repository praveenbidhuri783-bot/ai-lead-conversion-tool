"""Unit tests for Excel/CSV export and large-file scoring throughput."""

import json
import os
import time

import openpyxl
import pandas as pd
import pytest

from modules import exporter, scorer

CONFIG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")


@pytest.fixture(scope="module")
def scoring_config():
    with open(os.path.join(CONFIG_DIR, "scoring_config.json"), "r", encoding="utf-8") as f:
        return json.load(f)


@pytest.fixture(scope="module")
def pitch_config():
    with open(os.path.join(CONFIG_DIR, "pitch_config.json"), "r", encoding="utf-8") as f:
        return json.load(f)


@pytest.fixture(scope="module")
def scored_df(scoring_config, pitch_config):
    summaries = [
        "Customer is ready to book, wants the earliest slot.",
        "Not interested, please do not call again.",
        "Wrong number, this is not the right person.",
        "Customer requested a callback for tomorrow evening.",
        "Customer is unsure due to price concern.",
    ] * 4
    df = pd.DataFrame({"summary": summaries, "campaign_name": ["Camp A", "Camp B"] * 10})
    return scorer.score_dataframe(df, "summary", {"campaign_name": "campaign_name"}, scoring_config, pitch_config)


def test_excel_report_has_all_required_sheets(scored_df, scoring_config, pitch_config):
    buffer = exporter.build_excel_report(scored_df, {"campaign_name": "campaign_name"}, scoring_config, pitch_config)
    workbook = openpyxl.load_workbook(buffer)
    assert set(exporter.SHEET_ORDER) == set(workbook.sheetnames)


def test_all_lead_scoring_sheet_row_count(scored_df, scoring_config, pitch_config):
    buffer = exporter.build_excel_report(scored_df, {"campaign_name": "campaign_name"}, scoring_config, pitch_config)
    workbook = openpyxl.load_workbook(buffer)
    sheet = workbook["All Lead Scoring"]
    assert sheet.max_row - 1 == len(scored_df)


def test_csv_export_round_trips(scored_df):
    csv_bytes = exporter.build_csv_export(scored_df)
    assert csv_bytes.startswith(b"summary") or b"summary" in csv_bytes.splitlines()[0]


def test_priority_subset_excel(scored_df):
    buffer = exporter.build_priority_subset_excel(scored_df, ["P4 - Exclude"])
    workbook = openpyxl.load_workbook(buffer)
    sheet = workbook["Leads"]
    expected = int((scored_df["Lead_Priority"] == "P4 - Exclude").sum())
    assert sheet.max_row - 1 == expected


def test_large_file_processing(scoring_config, pitch_config):
    summaries = [
        "Customer wants to book, ready to proceed.",
        "Not interested, do not call.",
        "Customer requested a callback for tomorrow.",
    ] * 2000
    df = pd.DataFrame({"summary": summaries})

    start = time.time()
    scored = scorer.score_dataframe(df, "summary", {}, scoring_config, pitch_config)
    elapsed = time.time() - start

    assert len(scored) == 6000
    assert elapsed < 60
